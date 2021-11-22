import openpyxl

wb = openpyxl.load_workbook('example.xlsx')
print(wb.sheetnames)

sheet = wb['Sheet1']
#sheet = wb.get_sheet_by_name('Sheet1')

print(sheet.title)


for i in range(1, sheet.max_row+1):
    print(sheet.cell(row = i, column = 1).value, sheet.cell(row = i,column = 2).value, sheet.cell(row = i, column = 3).value)

"""
for row in sheet.rows:
    print(row[0].value, row[1].value, row[2].value)
"""

sheet['B1']='Cat'

wb.save('example.xlsx')
